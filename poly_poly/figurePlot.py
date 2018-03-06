'''This python script is used to process the general data from a file into
  line plot, and save the figure as a .png file'''
###########################################################################
# Import modules
from __future__ import division
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
from openpyxl import load_workbook

###########################################################################
# Data files:
# Import data from file
# .txt/.csv file
with open('poly_poly_devc.csv') as f1:
    lines = (line for line in f1 if not line.startswith('%'))
    ts, Tfs, Tms, Tbs = np.loadtxt(lines, unpack = True, delimiter=',', skiprows=2)

with open('Experiment.dat') as f2:
    lines = (line for line in f2 if not line.startswith('%'))
    te, Tfe, Tme1, Tme2, Tbe = np.loadtxt(lines, unpack = True, skiprows=0)

with open('matSim.dat') as f3:
    lines = (line for line in f3 if not line.startswith('%'))
    tm, Tfm, Tmm, Tbm = np.loadtxt(lines, unpack = True, skiprows=0)

# Import data from excelfile
# wb = load_workbook('surf_mass_part_char_cyl_fuel_devc.csv', read_only=True)
# sheet_1 = wb.get_sheet_by_name('Sheet1')
# t = np.zeros(sheet_1.max_row-1)
# V20M = np.zeros(sheet_1.max_row-1)

# for i in range(0,sheet_1.max_row-1):
#     t20M[i]=sheet_1.cell(row=i+2, column=1).value
#     V20M[i]=sheet_1.cell(row=i+2, column=2).value

# Manually input data
# RL = 15
# freq = np.int64(20)
# Width = np.array([600, 300, 150])
# Hp = np.array([[30, 50, 75, 100, 150, 200, 250, 300],
#               [30, 60, 90, 120, 150],
#                [25, 50, 75]])
# Up = np.array([[1.0823, 1.0913, 1.29, 1.4006, 1.33, 0.98, 0.46977, -0.1],
#               [0.236, -0.0558, -0.0179, 0.0065, -0.00498],
#                [-2.1387, -1.91, -0.0018]])

###########################################################################
# Custom greek font in label
plt.rc('font', family='serif')
plt.rc('font', serif='Times New Roman')
mpl.rcParams['mathtext.fontset'] = 'custom'
mpl.rcParams['mathtext.rm'] = 'Times New Roman'

# RGB color schemes from matlab
colors = [(0, 0.447, 0.7410), (0.8500, 0.3250, 0.0980),
            (0.9290, 0.6940, 0.1250)]

# Figure parameters
fsize = 20
hsize = 16
fwidth = 6.5
fheight = 5

###########################################################################
# Plot commenders
fig1 = plt.figure(figsize=(fwidth, fheight))
plt.plot(ts, Tfs, '-',color=colors[0], linewidth = 2, label = 'Sim_Front')
plt.plot(ts, Tbs, '--',color=colors[1], linewidth = 2, label = 'Sim_Back')
plt.plot(ts, Tms, '-.',color=colors[2], linewidth = 2, label = 'Sim_Middle')
plt.plot(te[1:-1:20], Tfe[1:-1:20]-273.35, 'k^', label = 'Exp_Front')
plt.plot(te[1:-1:20], Tbe[1:-1:20]-273.35, 'kv', label = 'Exp_Back')
plt.xlabel('Time (s)', fontname = 'Times New Roman', fontsize = fsize)
plt.ylabel(r'Temperature (C)', fontname = 'Times New Roman', fontsize = fsize)
H = plt.legend(loc = 'upper left', prop={'size':14}, numpoints = 1,
               frameon = False)
plt.rc('xtick', labelsize=16)
plt.rc('ytick', labelsize=16)
# plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
plt.axis([0, 200, 0, 500])

# Save figures
plt.tight_layout()
plt.savefig('FDS_T_distribution.png', dpi = 300)
plt.close()

###########################################################################
# Compare matlab and fds results in similar setup
fig2 = plt.figure(figsize=(fwidth, fheight))
plt.plot(ts, Tfs, '-',color=colors[0], linewidth = 2, label = 'FDS_Front')
plt.plot(ts, Tms, '-.',color=colors[2], linewidth = 2, label = 'FDS_Middle')
plt.plot(ts, Tbs, '--',color=colors[1], linewidth = 2, label = 'FDS_Back')
plt.plot(tm, Tfm-273, 'k^', label = 'Mat_Front')
plt.plot(tm, Tmm-273, 'ko', label = 'Mat_Middle')
plt.plot(tm, Tbm-273, 'kv', label = 'Mat_Back')
plt.xlabel('Time (s)', fontname = 'Times New Roman', fontsize = fsize)
plt.ylabel(r'Temperature (C)', fontname = 'Times New Roman', fontsize = fsize)
H = plt.legend(loc = 'upper left', prop={'size':14}, numpoints = 1,
               frameon = False)
plt.rc('xtick', labelsize=16)
plt.rc('ytick', labelsize=16)
# plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
plt.axis([0, 200, 0, 450])

plt.tight_layout()
plt.savefig('FDS_Matlab.png', dpi = 300)
plt.close()
